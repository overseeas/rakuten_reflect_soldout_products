from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time
import json
import datetime
import re
import os
import csv
import base64
import requests
from openpyxl import load_workbook
import shutil
import traceback
from skpy import Skype


#PATH
DOWNLOADS = os.path.abspath("output")
MASTER = "//acad2/Ace/020_制限共用部/520_システム/WinActor/物販本社/楽天欠品作業/マスタ.xlsx"
BACKUP = os.path.abspath("backup")

def get_credentials():
    f = open("config/vault.json")
    data = json.load(f)
    f.close()
    return data


def order_login(driver, credentials) -> None:
    driver.get("https://ace-1648.suruzo.biz/auth/login/?")
    if "株式会社エース" in driver.title:
        #id
        elem = driver.find_element(By.ID, "emp_cd")
        elem.clear()
        elem.send_keys(credentials["suruzo"]["id"])
        elem.send_keys(Keys.RETURN)

        #password
        elem = driver.find_element(By.ID, "password")
        elem.clear()
        elem.send_keys(credentials["suruzo"]["password"])
        elem.send_keys(Keys.RETURN)
        time.sleep(3)
        if "トップ画面" in driver.title:
            return True
    return False

def delete_files_in_directory(directory_path):
   try:
        with os.scandir(directory_path) as entries:
            for entry in entries:
                if entry.is_file():
                    os.unlink(entry.path)
        return True
   except OSError:
       return False

def order_search(driver, search_from, search_to):
    #在庫管理ページへ移動
    driver.find_element(By.LINK_TEXT, "商品・在庫管理").click()
    driver.find_element(By.LINK_TEXT, "在庫管理").click()
    assert driver.title == "【商品・在庫管理】 - 在庫管理"

    #検索期間from
    driver.find_element(By.NAME, "zaiko_updatetime_y").send_keys(search_from.year)
    driver.find_element(By.NAME, "zaiko_updatetime_m").send_keys(search_from.month)
    driver.find_element(By.NAME, "zaiko_updatetime_d").send_keys(search_from.day)
    driver.find_element(By.NAME, "zaiko_updatetime_hh").clear()
    driver.find_element(By.NAME, "zaiko_updatetime_hh").send_keys(search_from.hour)
    driver.find_element(By.NAME, "zaiko_updatetime_mm").clear()
    driver.find_element(By.NAME, "zaiko_updatetime_mm").send_keys(search_from.minute)
    
    #検索期間to
    driver.find_element(By.NAME, "zaiko_updatetime2_y").send_keys(search_to.year)
    driver.find_element(By.NAME, "zaiko_updatetime2_m").send_keys(search_to.month)
    driver.find_element(By.NAME, "zaiko_updatetime2_d").send_keys(search_to.day)
    driver.find_element(By.NAME, "zaiko_updatetime2_hh").clear()
    driver.find_element(By.NAME, "zaiko_updatetime2_hh").send_keys(search_to.hour)
    driver.find_element(By.NAME, "zaiko_updatetime2_mm").clear()
    driver.find_element(By.NAME, "zaiko_updatetime2_mm").send_keys(search_to.minute)
    
    #検索
    driver.find_element(By.NAME, "normal_search").click()
    search_result = driver.find_element(By.CLASS_NAME, "headPageChanger").text
    
    #検索結果の数を返し
    counts = re.search("[,\d]+(?=件)", search_result)
    return counts.group(0)
    
def download_file(driver, path_to_downloads, extension):
    driver.find_element(By.NAME, "output_p_main").click()
    seconds = 0
    dl_wait = False
    #60秒内にダウンロードされるか確認
    while not(dl_wait) and seconds < 60:
        time.sleep(1)
        dl_wait = False
        for fname in os.listdir(path_to_downloads):
            if fname.endswith('.' + extension):
                dl_wait = True
        seconds += 1
    return path_to_downloads + "\\" + fname


def get_search_period():
    #検索期間_to
    period_to = datetime.datetime.now()
    
    #検索期間_from
    f = open("config/latestTime.txt")
    str_period_from = f.read()
    f.close()
    period_from = datetime.datetime.strptime(str_period_from, "%Y-%m-%d %H:%M:%S")
    return([period_from, period_to])

def open_browser():
    delete_files_in_directory(DOWNLOADS)
    chromeOptions = webdriver.ChromeOptions()
    prefs = {"download.default_directory" : DOWNLOADS}
    chromeOptions.add_experimental_option("prefs",prefs)
    chromeOptions.add_argument('--headless')
    service = Service(excutable_path = "C:\\Users\\winact_user\\Documents\\WinActor\\webdriver\\chromedriver.exe")
    return webdriver.Chrome(service=service, options=chromeOptions)

def verify_with_master(data):
    wb = load_workbook(MASTER, read_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if data == cell.value:
                return ws.cell(row=cell.row, column=20).value
    return None

def encode_api_credentials(service_secret, licenseKey):
    data = service_secret + ":" + licenseKey
    byte_data = data.encode('utf-8')
    return {"Authorization" : "ESA " + base64.b64encode(byte_data).decode("utf-8")}

def get_sku(managenumber, color_code, power):
    #度数0.00の場合
    if power == "0.00":
        power = "±0.00(度なし)"
    credentials = get_credentials()
    auth_headers = encode_api_credentials(credentials["rakuten"]["serviceSecret"], credentials["rakuten"]["licenseKey"])
    r = requests.get("https://api.rms.rakuten.co.jp/es/2.0/items/manage-numbers/" + str(managenumber), headers=auth_headers)
    time.sleep(0.2)
    if r.status_code == 200:
        try:
            data = r.json()["variants"]
            for sku, code in data.items():
                if "(" + color_code + ")" in code["selectorValues"]["Key0"] and code["selectorValues"]["Key1"] == power:
                    return sku
        except:
            pass
    return False


def update_stock(managenumber, sku, quantity):
    credentials = get_credentials()
    auth_headers = encode_api_credentials(credentials["rakuten"]["serviceSecret"], credentials["rakuten"]["licenseKey"])
    auth_headers.update({"Content-Type": "application/json"})
    r = requests.put(f"https://api.rms.rakuten.co.jp/es/2.0/inventories/manage-numbers/{managenumber}/variants/{sku}", json={"mode": "ABSOLUTE","quantity": quantity}, headers=auth_headers)
    if r.status_code != 204:
        return False
    r = requests.get(f"https://api.rms.rakuten.co.jp/es/2.0/inventories/manage-numbers/{managenumber}/variants/{sku}", headers=auth_headers)
    if r.status_code == 200:
        data = r.json()
        if data["quantity"] != quantity:
            return False
    else:
        return False
    time.sleep(1)
    return True
        
def record_searched_time(time):
    with open("config/latestTime.txt", "w", encoding='utf-8') as f:
        f.write(time.strftime("%Y-%m-%d %H:%M:%S"))

def backup_data(data, time, backupdata):
    path = os.path.join(BACKUP, time.strftime("%Y%m%d%H%M"))
    os.makedirs(path)
    shutil.copy2(data, path)
    with open(path + '/upload_body.csv', 'w', newline='', encoding='utf-8-sig') as f:
        spamwriter = csv.writer(f, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
        spamwriter.writerow(["商品管理番号（商品URL）", "カラー", "度数", "在庫数"])
        spamwriter.writerows(backupdata)


def skype_send(live_id, message):
    credentials = get_credentials()
    sk = Skype(credentials["skype"]["id"], credentials["skype"]["password"])

    # 送信先の設定
    ch = sk.contacts[live_id].chat
    ch.sendMsg(message, rich=True)

def main():
    #initiate
    bulk = []

    credentials = get_credentials()
    search_period = get_search_period()
    driver = open_browser()
    if not(order_login(driver, credentials)):
        skype_send(credentials["oota"]["skypeLiveId"], "通販する蔵ログイン失敗しました。</br>楽天欠品作業を中止します。")
        return False
    orders_num = order_search(driver, search_period[0], search_period[1])
    #orders_num = 1
    #input("pause")
    if orders_num == "0":
        print("0件")
    else:
        downloaded = download_file(driver, DOWNLOADS, "csv")
        
        with open(downloaded, "r", newline="", encoding="shift_jis") as csvfile:
            stock_info = csv.DictReader(csvfile)

            tomin = []
            tomax = []
            
            #全行を繰り返して情報を取得し、API用データを作成
            for row in stock_info:
                
                managenumber = re.sub("'", "", row["自社品番"])
                color = re.findall("(?<=\().+(?=\))", re.sub("'", "", row["カラー"]))[0]
                power = re.sub("'", "", row["サイズ"])
                stock = re.sub("'", "", row["サイト在庫数"])
                status = re.sub("'", "",row["メーカー在庫"])

                #マスターファイルにあるか確認して、T列の値を取得
                minimum_stock = verify_with_master(color)
                sku = get_sku(managenumber, color, power)
                
                #bulkデータを作成
                if sku and minimum_stock != None:
                    if power == '0.00':
                        minimum_stock *= 3
                    #欠品処理
                    if status == "終売" or (status == "欠品" and int(stock) < minimum_stock):
                        quantity = 0
                    #欠品解消処理
                    else:
                        quantity = 9999
                    if not(update_stock(managenumber, sku, quantity)):
                        skype_send(credentials["oota"]["skypeLiveId"], "楽天欠品処理に失敗しました。</br>backupフォルダを確認してください。")
                        return False
                    
                    #backupデータの作成
                    backupinfo = [managenumber, re.sub("'", "", row["カラー"]), power, quantity]
                    if quantity == 0:
                        tomin.append(backupinfo)
                    else:
                        tomax.append(backupinfo)
        backup_data(downloaded, search_period[1], tomin + tomax)
        delete_files_in_directory(DOWNLOADS)
    record_searched_time(search_period[1])

    


if __name__ == "__main__":
    try:
        main()
    except:
        path = os.path.join(BACKUP, datetime.datetime.now().strftime("%Y%m%d%H%M"))
        os.makedirs(path)
        with open(os.path.join(path, "error.log"), 'w', encoding='utf-8') as f:
            traceback.print_exc(file=f)